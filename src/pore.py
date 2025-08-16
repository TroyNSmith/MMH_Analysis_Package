import click
import MDAnalysis as mda
import mdtraj as mdt
import numpy as np
import os
import pandas as pd
import sys

from openpyxl import Workbook
from pathlib import Path
sys.path.append(str(Path.cwd()))

import mdevaluate as mde
from functions import mdevaluate_analysis, mdtraj_analysis
from utils.coordinates import centers_of_masses, vectorize_residue
from utils.plotting import plot_heatmap, plot_line
from functions import mdanalysis_analysis as hbonds
from utils.logging import log_analysis_yaml


@click.group()
def cli():
    """Main CLI entry point."""
    pass


@click.command()
@click.option(
    "-i",
    "--input",
    "sim_dir",
    help="Path to simulation directory containing trajectory and topology files. This directory will also be used to dump output files unless otherwise specified.",
)
@click.option(
    "-o",
    "--output",
    "out_dir",
    help="Path to dump output files from analysis.",
    default=None,
)
@click.option(
    "-tr",
    "--trajectory",
    "trajectory",
    help="Name of the trajectory file located within -d / --directory.",
)
@click.option(
    "-tp",
    "--topology",
    "topology",
    help="Name of the topology file located within -d / --directory.",
)
@click.option(
    "-st",
    "--structure",
    "structure",
    help="Name of the structure (i.e. .gro) file located within -d / --directory.",
)
@click.option(
    "-r",
    "--resname",
    "res_name",
    help="Name of the residue to be analyzed.",
    default=None,
)
@click.option(
    "-a",
    "--atoms",
    "atoms",
    nargs=2,
    help="Two atoms for analysis located on the residue specified in -res / --resname.",
    default=None,
)
@click.option(
    "-s",
    "--segments",
    "num_segments",
    help="Number of starting points to use with correlation functions.",
    type=click.IntRange(10, 1000, clamp=True),
    default=500,
)
@click.option(
    "-d",
    "--diameter",
    "pore_diameter",
    help="Diameter of the pore in nm.",
    type=click.FloatRange(0, 1000, clamp=False),
    default=None,
)
@click.option(
    "-q",
    "--q",
    "q_val",
    type=float,
    help="Magnitude of the scattering vector, q.",
    default=None,
)
@click.option(
    "-ov",
    "--override",
    "override",
    help="Force re-run of calculations.",
    is_flag=True,
)
@click.option(
    "-po",
    "--plot_only",
    "plot_only",
    help="Skip calculations and regenerate plots of existing data.",
    is_flag=True,
)
def Run(
    sim_dir: Path,
    out_dir: Path,
    trajectory: str,
    topology: str,
    structure: str,
    res_name: str,
    atoms: list[str],
    num_segments: int,
    pore_diameter: float,
    q_val: float,
    override: bool,
    plot_only: bool,
):
    # ==============================
    # ===== Step 1: Initialize =====
    # ==============================
    mde_coords = mde.open(
        directory=sim_dir, topology=topology, trajectory=trajectory, nojump=True
    )
    mde_vectors = vectorize_residue(
        mde_coords=mde_coords, res_name=res_name, atoms=atoms
    )
    mde_com = centers_of_masses(mde_coords=mde_coords, res_name=res_name)

    mde_atom_1_coords = mde_coords.subset(atom_name=atoms[0], residue_name=res_name)

    mde_atom_2_coords = mde_coords.subset(atom_name=atoms[1], residue_name=res_name)

    mda_coords = mda.Universe(
        os.path.join(sim_dir, topology), os.path.join(sim_dir, trajectory)
    )

    mdt_coords = mdt.load(
        os.path.join(sim_dir, trajectory), top=os.path.join(sim_dir, structure)
    )

    mdt_resid_idxs = mdt_coords.topology.select(f"resname {res_name}")
    mdt_resid_traj = mdt_coords.atom_slice(mdt_resid_idxs)

    dir_out = (
        os.path.join(out_dir, "Analysis", f"{res_name}_{atoms[0]}_ {atoms[1]}")
        if out_dir is not None
        else os.path.join(sim_dir, "Analysis", f"{res_name}_{atoms[0]}_{atoms[1]}")
    )

    os.makedirs(dir_out, exist_ok=True)

    yaml_out = Path(os.path.join(dir_out, "log.yaml"))
    xlsx_out = Path(
        os.path.join(dir_out, f"results_{res_name}_{atoms[0]}_{atoms[1]}.xlsx")
    )

    params = {
        "sim_directory": sim_dir,
        "trajectory": trajectory,
        "topology": topology,
        "res_name": res_name,
        "atoms": atoms,
        "num_segments": num_segments,
        "pore_diameter": pore_diameter,
        "q_magnitude": q_val,
    }

    if not os.path.exists(xlsx_out):
        wb = Workbook()
        ws = wb["Sheet"]
        ws.title = "Parameters"
        ws.cell(
            row=1,
            column=1,
            value="Check summary.yaml for more accurate parameters as those listed here are for the initial analysis only (i.e., not necessarily for regenerated analyses.)",
        )
        for i, (param, value) in enumerate(params.items(), start=2):
            ws.cell(row=i, column=1, value=param)
            ws.cell(row=i, column=2, value=str(value))
        wb.save(xlsx_out)

    # =================================================
    # ===== Step 2: Radial distribution functions =====
    # =================================================
    os.makedirs(os.path.join(dir_out, "RDF"), exist_ok=True)

    # ===== Total (centers of masses)
    csv_out_total = os.path.join(dir_out, "RDF", "RDF_total.csv")
    if not os.path.exists(csv_out_total) or override:
        print("Calculating total centers of masses RDF...")
        rdf_total = mdevaluate_analysis.radial_distribution_function(
            coords_1=mde_com,
            num_segments=num_segments,
            column_label="Centers of Masses",
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Centers of Masses RDF (total)",
            file_path=csv_out_total,
            parameters=params,
        )
        rdf_total.to_csv(csv_out_total, index=False)

    # ===== Magnitude of q vector
    if q_val is None:
        rdf_total = pd.read_csv(csv_out_total)
        y_max_idx = rdf_total.iloc[:, 1].idxmax()
        x_at_max_y = rdf_total.iloc[y_max_idx, 0]
        q_val = float(round(2 * np.pi / x_at_max_y, 3))
        params["q_magnitude"] = f"{q_val} (auto-generated)"

    # ===== Intramolecular
    csv_out_intra = os.path.join(dir_out, "RDF", "RDF_intra.csv")
    if not os.path.exists(csv_out_intra) or override:
        print("Calculating intramolecular RDF...")
        rdf_intra = mdevaluate_analysis.radial_distribution_function(
            coords_1=mde_atom_1_coords,
            coords_2=mde_atom_2_coords,
            num_segments=num_segments,
            mode="intra",
            column_label=f"Intra {atoms[0]}:{atoms[1]}",
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Intra RDF",
            file_path=csv_out_intra,
            parameters=params,
        )
        rdf_intra.to_csv(csv_out_intra, index=False)

    # ===== Intermolecular (atom 1)
    csv_out_inter_1 = os.path.join(dir_out, "RDF", f"RDF_inter_{atoms[0]}.csv")
    if not os.path.exists(csv_out_inter_1) or override:
        print("Calculating intermolecular RDF #1...")
        rdf_inter_1 = mdevaluate_analysis.radial_distribution_function(
            coords_1=mde_atom_1_coords,
            coords_2=mde_atom_1_coords,
            num_segments=num_segments,
            mode="inter",
            column_label=f"Inter {atoms[0]}:{atoms[0]}",
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name=f"Inter RDF ({atoms[0]})",
            file_path=csv_out_inter_1,
            parameters=params,
        )
        rdf_inter_1.to_csv(csv_out_inter_1, index=False)

    # ===== Intermolecular (atom 2)
    csv_out_inter_2 = os.path.join(dir_out, "RDF", f"RDF_inter_{atoms[1]}.csv")
    if not os.path.exists(csv_out_inter_2) or override:
        print("Calculating intermolecular RDF #2...")
        rdf_inter_2 = mdevaluate_analysis.radial_distribution_function(
            coords_1=mde_atom_2_coords,
            coords_2=mde_atom_2_coords,
            num_segments=num_segments,
            mode="inter",
            column_label=f"Inter {atoms[1]}:{atoms[1]}",
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name=f"Inter RDF ({atoms[1]})",
            file_path=csv_out_inter_2,
            parameters=params,
        )
        rdf_inter_2.to_csv(csv_out_inter_2, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "RDF", "RDF.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            rdf_total = pd.read_csv(csv_out_total)
            rdf_intra = pd.read_csv(csv_out_intra)
            rdf_inter_1 = pd.read_csv(csv_out_inter_1)
            rdf_inter_2 = pd.read_csv(csv_out_inter_2)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        msd_merged = pd.concat(
            [
                rdf_total,
                rdf_intra.iloc[:, 1],
                rdf_inter_1.iloc[:, 1],
                rdf_inter_2.iloc[:, 1],
            ],
            axis=1,
        )

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            msd_merged.to_excel(writer, sheet_name="RDF", index=False)

        plot_line(
            output_path=png_out,
            x_data=msd_merged.iloc[:, 0],
            y_data=msd_merged.iloc[:, 1:],
            x_axis_label="r / nm",
            y_axis_label="g(r)",
        )

    # ============================================
    # ===== Step 3: Mean square displacement =====
    # ============================================
    os.makedirs(os.path.join(dir_out, "MSD"), exist_ok=True)

    # ===== Residual MSD (all directions)
    csv_out_all = os.path.join(dir_out, "MSD", "MSD_all.csv")
    if not os.path.exists(csv_out_all) or override:
        print("Calculating total MSD...")
        msd_all = mdevaluate_analysis.mean_square_displacement(
            coords=mde_com, num_segments=num_segments
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="MSD (all)",
            file_path=csv_out_all,
            parameters=params,
        )
        msd_all.to_csv(csv_out_all, index=False)

    # ===== Residual MSD (z-axis)
    csv_out_z = os.path.join(dir_out, "MSD", "MSD_z.csv")
    if not os.path.exists(csv_out_z) or override:
        print("Calculating z-axis MSD...")
        msd_z = mdevaluate_analysis.mean_square_displacement(
            coords=mde_com, axis="z", num_segments=num_segments
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="MSD (z-axis)",
            file_path=csv_out_z,
            parameters=params,
        )
        msd_z.to_csv(csv_out_z, index=False)

    # ===== Residual MSD (xy-plane)
    csv_out_xy = os.path.join(dir_out, "MSD", "MSD_xy.csv")
    if not os.path.exists(csv_out_xy) or override:
        print("Calculating xy-plane MSD...")
        msd_xy = mdevaluate_analysis.mean_square_displacement(
            coords=mde_com, axis="xy", num_segments=num_segments
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="MSD (xy-plane)",
            file_path=csv_out_xy,
            parameters=params,
        )
        msd_xy.to_csv(csv_out_xy, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "MSD", "MSD.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            msd_all = pd.read_csv(csv_out_all)
            msd_z = pd.read_csv(csv_out_z)
            msd_xy = pd.read_csv(csv_out_xy)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        msd_merged = pd.concat([msd_all, msd_z.iloc[:, 1], msd_xy.iloc[:, 1]], axis=1)

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            msd_merged.to_excel(writer, sheet_name="MSD", index=False)

        plot_line(
            output_path=png_out,
            x_data=msd_merged.iloc[:, 0],
            y_data=msd_merged.iloc[:, 1:],
            x_axis_scale="log",
            x_axis_label=r"$\mathbf{\mathit{t}}$ / ps",
            y_axis_scale="log",
            y_axis_label=r"<r$^2$> / nm$^2$$\cdot$ps$^{-1}$",
        )

    # ==============================================================
    # ===== Step 4: Radially resolved mean square displacement =====
    # ==============================================================
    csv_out_resolved = os.path.join(dir_out, "MSD", "MSD_resolved.csv")
    if not os.path.exists(csv_out_resolved) or override:
        print("Calculating radially resolved MSD...")
        msd_resolved = mdevaluate_analysis.mean_square_displacement(
            coords=mde_com,
            num_segments=num_segments,
            radially_resolved=True,
            pore_diameter=pore_diameter,
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="MSD (radially resolved)",
            file_path=csv_out_resolved,
            parameters=params,
        )
        msd_resolved.to_csv(csv_out_resolved, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "MSD", "MSD_resolved.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            msd_resolved = pd.read_csv(csv_out_resolved)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            msd_resolved.to_excel(writer, sheet_name="MSD Resolved", index=False)

        plot_line(
            output_path=png_out,
            x_data=msd_resolved.iloc[:, 0],
            y_data=msd_resolved.iloc[:, 1:],
            x_axis_scale="log",
            x_axis_label=r"$\mathbf{\mathit{t}}$ / ps",
            y_axis_scale="log",
            y_axis_label=r"<r$^2$> / nm$^2$$\cdot$ps$^{-1}$",
        )

    # ==================================================
    # ===== Step 5: Incoherent scattering function =====
    # ==================================================
    os.makedirs(os.path.join(dir_out, "ISF"), exist_ok=True)

    # ===== Residual ISF (total)
    csv_out_isf = os.path.join(dir_out, "ISF", "ISF_total.csv")
    if not os.path.exists(csv_out_isf) or override:
        print("Calculating ISF...")
        isf_total = mdevaluate_analysis.incoherent_scattering_function(
            coords=mde_com,
            q_val=q_val,
            num_segments=num_segments,
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="ISF (total)",
            file_path=csv_out_isf,
            parameters=params,
        )
        isf_total.to_csv(csv_out_isf, index=False)

    # ===== Residual ISF (xy-plane)
    csv_out_isf_xy = os.path.join(dir_out, "ISF", "ISF_xy_plane.csv")
    if not os.path.exists(csv_out_isf_xy) or override:
        print("Calculating ISF...")
        isf_total = mdevaluate_analysis.incoherent_scattering_function(
            coords=mde_com,
            q_val=q_val,
            num_segments=num_segments,
            axis="xy",
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="ISF (xy-plane)",
            file_path=csv_out_isf_xy,
            parameters=params,
        )
        isf_total.to_csv(csv_out_isf_xy, index=False)

    # ===== Residual ISF (z-axis)
    csv_out_isf_z = os.path.join(dir_out, "ISF", "ISF_z_axis.csv")
    if not os.path.exists(csv_out_isf_z) or override:
        print("Calculating ISF...")
        isf_total = mdevaluate_analysis.incoherent_scattering_function(
            coords=mde_com,
            q_val=q_val,
            num_segments=num_segments,
            axis="z",
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="ISF (z-axis)",
            file_path=csv_out_isf_z,
            parameters=params,
        )
        isf_total.to_csv(csv_out_isf_z, index=False)

    # ===== Residual ISF (resolved)
    csv_out_isf_resolved = os.path.join(dir_out, "ISF", "ISF_resolved.csv")
    if not os.path.exists(csv_out_isf_resolved) or override:
        print("Calculating resolved ISF...")
        isf_resolved = mdevaluate_analysis.incoherent_scattering_function(
            coords=mde_com,
            q_val=q_val,
            num_segments=num_segments,
            radially_resolved=True,
            pore_diameter=pore_diameter,
            num_bins=3,
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="ISF (resolved)",
            file_path=csv_out_isf_resolved,
            parameters=params,
        )
        isf_resolved.to_csv(csv_out_isf_resolved, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "ISF", "ISF.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            isf_total = pd.read_csv(csv_out_isf)
            isf_xy_plane = pd.read_csv(csv_out_isf_xy)
            isf_z_axis = pd.read_csv(csv_out_isf_z)
            isf_resolved = pd.read_csv(csv_out_isf_resolved)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        isf_merged = pd.concat(
            [
                isf_total,
                isf_xy_plane.iloc[:, 1:],
                isf_z_axis.iloc[:, 1:],
                isf_resolved.iloc[:, 1:],
            ],
            axis=1,
        )

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            isf_merged.to_excel(writer, sheet_name="ISF", index=False)

        plot_line(
            output_path=png_out,
            x_data=isf_merged.iloc[:, 0],
            y_data=isf_merged.iloc[:, 1:],
            x_axis_scale="log",
            x_axis_label="t / ps",
            y_axis_label=f"ISF(q={q_val:.1f}, t)",
        )

    # ==================================================
    # ===== Step 6: Rotational correlation coeffs. =====
    # ==================================================
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Residual ISF
    csv_out_rotat_correl = os.path.join(
        dir_out, "Etc", "Rotational_correlation_coeffs.csv"
    )
    if not os.path.exists(csv_out_rotat_correl) or override:
        print("Calculating rotational correlation coefficients...")
        rotat_correl = mdevaluate_analysis.rotational_correlations(
            vectors=mde_vectors, num_segments=num_segments
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Rotational correlation coefficients",
            file_path=csv_out_rotat_correl,
            parameters=params,
        )
        rotat_correl.to_csv(csv_out_rotat_correl, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Rotational_correlation_coeffs.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            rotat_correl = pd.read_csv(csv_out_rotat_correl)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            rotat_correl.to_excel(
                writer, sheet_name="Rotational correlation coefficients", index=False
            )

        plot_line(
            output_path=png_out,
            x_data=rotat_correl.iloc[:, 0],
            y_data=rotat_correl.iloc[:, 1:],
            x_axis_label="t / ps",
            x_axis_scale="log",
            y_axis_label=r"$F_n(t)$",
        )

    # ========================================================
    # ===== Step 7: Rotational van Hove Self-Correlation =====
    # ========================================================
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Rotational van Hove
    csv_out_rotat_van_Hove = os.path.join(dir_out, "Etc", "Rotational_van_Hove.csv")
    if not os.path.exists(csv_out_rotat_van_Hove) or override:
        print("Calculating rotational van Hove self-correlation...")
        rotat_van_Hove = mdevaluate_analysis.van_hove_rotation(
            vectors=mde_vectors, num_segments=num_segments
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Rotational van Hove",
            file_path=csv_out_rotat_van_Hove,
            parameters=params,
        )
        rotat_van_Hove.to_csv(csv_out_rotat_van_Hove, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Rotational_van_Hove.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            rotat_van_Hove = pd.read_csv(csv_out_rotat_van_Hove)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            rotat_van_Hove.to_excel(
                writer, sheet_name="Rotational van Hove", index=False
            )

        plot_line(
            output_path=png_out,
            x_data=rotat_van_Hove.iloc[:, 0],
            y_data=rotat_van_Hove.iloc[:, 2:],
            x_axis_label="θ / degrees",
            y_axis_label="F(θ, t)",
        )

    # ===========================================================
    # ===== Step 8: Translational van Hove Self-Correlation =====
    # ===========================================================
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Translational van Hove
    csv_out_trans_van_Hove = os.path.join(dir_out, "Etc", "Translational_van_Hove.csv")
    if not os.path.exists(csv_out_trans_van_Hove) or override:
        print("Calculating translational van Hove self-correlation...")
        trans_van_Hove = mdevaluate_analysis.van_hove_translation(
            coords=mde_com,
            num_segments=num_segments,
            pore_diameter=pore_diameter,
            num_bins=250,
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Translational van Hove",
            file_path=csv_out_trans_van_Hove,
            parameters=params,
        )
        trans_van_Hove.to_csv(csv_out_trans_van_Hove, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Translational_van_Hove.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            trans_van_Hove = pd.read_csv(csv_out_trans_van_Hove)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            trans_van_Hove.to_excel(
                writer, sheet_name="Translational van Hove", index=False
            )

        plot_line(
            output_path=png_out,
            x_data=trans_van_Hove.iloc[:, 0],
            y_data=trans_van_Hove.iloc[:, 2:],
            x_axis_label="r / nm",
            y_axis_label="F(r, t)",
        )

    # ==========================================
    # ===== Step 9: Non-Gaussian Parameter =====
    # ==========================================
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Non-Gaussian alpha parameter
    csv_out_non_gauss = os.path.join(dir_out, "Etc", "non_Gaussian_parameter.csv")
    if not os.path.exists(csv_out_non_gauss) or override:
        print("Calculating non-Gaussian parameters...")
        non_gauss_param = mdevaluate_analysis.non_gaussian_parameter(
            coords=mde_com,
            num_segments=num_segments,
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Non-Gaussian parameter",
            file_path=csv_out_non_gauss,
            parameters=params,
        )
        non_gauss_param.to_csv(csv_out_non_gauss, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "non_Gaussian_parameter.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            non_gauss_param = pd.read_csv(csv_out_non_gauss)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            non_gauss_param.to_excel(
                writer, sheet_name="Non-Gaussian Parameter", index=False
            )

        plot_line(
            output_path=png_out,
            x_data=non_gauss_param.iloc[:, 0],
            y_data=non_gauss_param.iloc[:, 1:],
            x_axis_label="t / ns",
            x_axis_scale="log",
            y_axis_label="NGP(t)",
        )

    # ==========================================
    # ===== Step 10: Chi-4 Susceptibility ======
    # ==========================================
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Chi-4 Susceptibility
    csv_out_chi_4 = os.path.join(dir_out, "Etc", "Chi_4_susceptibility.csv")
    if not os.path.exists(csv_out_chi_4) or override:
        print("Calculating chi-4 susceptibility...")
        chi_4 = mdevaluate_analysis.chi_4_susceptibility(
            coords=mde_com,
            q_val=q_val,
            num_segments=num_segments,
        )
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Chi-4 susceptibility",
            file_path=csv_out_chi_4,
            parameters=params,
        )
        chi_4.to_csv(csv_out_chi_4, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Chi_4_susceptibility.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            chi_4 = pd.read_csv(csv_out_chi_4)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            chi_4.to_excel(writer, sheet_name="Chi-4 Susceptibility", index=False)

        plot_line(
            output_path=png_out,
            x_data=chi_4.iloc[:, 0],
            y_data=chi_4.iloc[:, 1:],
            x_axis_label="t / ns",
            x_axis_scale="log",
            y_axis_label=rf"$χ_4(q={q_val:.1f}, t)$",
        )

    # =============================================
    # ===== Step 11: Spatial density function =====
    # =============================================
    os.makedirs(os.path.join(dir_out, "RDF"), exist_ok=True)

    # ===== Spatial density function
    csv_out_sdf = os.path.join(dir_out, "RDF", "Spatial_density.csv")
    if not os.path.exists(csv_out_sdf) or override:
        print("Calculating spatial density functions...")

        res_atom_pairs = {
            res_name: atoms,
            "LNK": ["NL"],
            "ETH": ["OEE"],
            "VAN": ["NV", "OVE", "OVH"],
        }

        sdf = mdevaluate_analysis.spatial_density_function(
            coords=mde_coords,
            res_atom_pairs=res_atom_pairs,
            pore_diameter=pore_diameter,
        )
        params["res_atom_pairs"] = res_atom_pairs
        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Spatial density function",
            file_path=csv_out_sdf,
            parameters=params,
        )
        del params["res_atom_pairs"]
        sdf.to_csv(csv_out_sdf, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "RDF", "Spatial_density.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            sdf = pd.read_csv(csv_out_sdf)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            sdf.to_excel(writer, sheet_name="Spatial density function", index=False)

        plot_line(
            output_path=png_out,
            x_data=sdf.iloc[:, 0],
            y_data=sdf.iloc[:, 1:],
            x_axis_label="r / nm",
            y_axis_label=r"Number density / Units $\cdot$ nm$^3$",
        )

    # ==================================================
    # ===== Step 12: Z-axis and xy-plane alignment =====
    # ==================================================
    """
    For best results, utilize (small) vectorized molecular segments (e.g. OCT H00: O01) rather than whole vectorized residues (e.g. OCT O01: C0O)
    """
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Z-axis alignment
    csv_out_z_axis_alignment = os.path.join(dir_out, "Etc", "Z_axis_alignment.csv")
    if not os.path.exists(csv_out_z_axis_alignment) or override:
        print("Calculating z-axis alignment...")
        z_axis_alignment = mdevaluate_analysis.ref_vector_alignment(
            vectors=mde_vectors, ref_vector=[0, 0, 1], num_segments=num_segments
        )

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Z-axis alignment",
            file_path=csv_out_z_axis_alignment,
            parameters=params,
        )
        z_axis_alignment.to_csv(csv_out_z_axis_alignment, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Z_axis_alignment.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            z_axis_alignment = pd.read_csv(csv_out_z_axis_alignment)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            z_axis_alignment.to_excel(
                writer, sheet_name="Z-axis Alignment", index=False
            )

        plot_line(
            output_path=png_out,
            x_data=z_axis_alignment.iloc[:, 0],
            y_data=z_axis_alignment.iloc[:, 1:],
            x_axis_label=r"$\theta$ / degrees",
            y_axis_label=r"$\alpha_{\vec{z}}$($\theta$, t)",
        )

    # ===== xy-plane alignment
    csv_out_xy_plane_alignment = os.path.join(dir_out, "Etc", "xy_plane_alignment.csv")
    if not os.path.exists(csv_out_xy_plane_alignment) or override:
        print("Calculating xy-plane alignment...")
        z_axis_alignment = mdevaluate_analysis.ref_plane_alignment(
            vectors=mde_vectors, normal_vector=[0, 0, 1], num_segments=num_segments
        )

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="xy-plane alignment",
            file_path=csv_out_xy_plane_alignment,
            parameters=params,
        )
        z_axis_alignment.to_csv(csv_out_xy_plane_alignment, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "xy_plane_alignment.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            xy_plane_alignment = pd.read_csv(csv_out_xy_plane_alignment)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            xy_plane_alignment.to_excel(
                writer, sheet_name="xy-Plane Alignment", index=False
            )

        plot_line(
            output_path=png_out,
            x_data=xy_plane_alignment.iloc[:, 0],
            y_data=xy_plane_alignment.iloc[:, 1:],
            x_axis_label=r"$\theta$ / degrees",
            y_axis_label=r"$\alpha_{XY}$($\theta$, t)",
        )

    # ============================================
    # ===== Step 13: Identify hydrogen bonds =====
    # ============================================
    os.makedirs(os.path.join(dir_out, "HBonds"), exist_ok=True)

    # ===== Initialize hydrogen bonds
    csv_out_h_bonds = os.path.join(dir_out, "HBonds", "HBonds.csv")
    if not os.path.exists(csv_out_h_bonds) or override:
        print("Identifying hydrogen bonds...")

        h_bonds_df, mda_hbonds = hbonds.identify_h_bonds(
            mda_coords=mda_coords, res_name=res_name
        )

        h_bonds_df.to_csv(csv_out_h_bonds, index=False)

        try:
            with pd.ExcelWriter(
                xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                h_bonds_df.to_excel(writer, sheet_name="H Bonds", index=False)
        except ValueError:
            pass

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Initialized hydrogen bonds",
            file_path=csv_out_h_bonds,
            parameters=params,
        )

    # ===== Analyze hydrogen bonds
    xlsx_out_hbonds = Path(os.path.join(dir_out, "HBonds", "hbonds_analysis.xlsx"))
    if not os.path.exists(xlsx_out_hbonds) or override or True is True:
        try:
            h_bonds_df = pd.read_csv(csv_out_h_bonds)

        except FileNotFoundError:
            raise FileNotFoundError(
                "Hydrogen bonds must be saved to .csv before analyzing."
            )
        print("Analyzing hydrogen bonds...")
        wb = Workbook()
        wb.save(xlsx_out_hbonds)

        results = hbonds.analyze_h_bonds(
            h_bonds_df=h_bonds_df, xlsx_out=xlsx_out_hbonds
        )

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Hydrogen bond analysis",
            file_path=str(xlsx_out_hbonds),
            parameters=params,
            results=results,
        )

    # ===== Hydrogen bonds heatmap
    png_out = os.path.join(dir_out, "HBonds", "All_h_bonds_heatmap.png")
    if not os.path.exists(png_out) or override:
        try:
            h_bonds_df = pd.read_csv(csv_out_h_bonds)

        except FileNotFoundError:
            raise FileNotFoundError(
                "Hydrogen bonds must be saved to .csv before analyzing."
            )

        print("Generating hydrogen bond heatmaps...")

        x_mesh, y_mesh, heatmaps = hbonds.hbonds_heatmap(
            mda_coords=mda_coords, h_bonds_df=h_bonds_df, pore_diameter=pore_diameter
        )

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Hydrogen bond heatmap",
            file_path=png_out,
            parameters=params,
        )

        for pair, heatmap in heatmaps.items():
            png_out = os.path.join(dir_out, "HBonds", f"{pair}_heatmap.png")
            plot_heatmap(
                output_path=png_out, x_mesh=x_mesh, y_mesh=y_mesh, heatmap=heatmap
            )

    # ===== Hydrogen bond clusters
    csv_out_clusters = os.path.join(dir_out, "HBonds", "H_bond_clusters.csv")
    # if not os.path.exists(csv_out_clusters) or override:
    if True is False:  # Temporary skip (taking too long for testing)
        try:
            h_bonds_df = pd.read_csv(csv_out_h_bonds)

        except FileNotFoundError:
            raise FileNotFoundError(
                "Hydrogen bonds must be saved to .csv before analyzing."
            )

        print("Identifying hydrogen bond clusters...")

        clusters, stats = hbonds.find_clusters(h_bonds_df=h_bonds_df)

        clusters.to_csv(csv_out_clusters, index=False)

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Hydrogen bond clusters",
            file_path=png_out,
            parameters=params,
            results=stats,
        )

    # ===========================================================
    # ===== Step 14: Radius of gyration and gyration tensor =====
    # ===========================================================
    os.makedirs(os.path.join(dir_out, "Etc"), exist_ok=True)

    # ===== Radius of gyration
    csv_out_rg = os.path.join(dir_out, "Etc", "Radius_of_gyration.csv")
    if not os.path.exists(csv_out_rg) or override:
        print("Calculating radii of gyration...")

        rad_gyration = mdtraj_analysis.radius_of_gyration(mdt_coords=mdt_resid_traj)

        results = {"Radius of gyration (average)": rad_gyration["Rg(t)"].mean()}

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Radius of gyration",
            file_path=csv_out_rg,
            parameters=params,
        )
        rad_gyration.to_csv(csv_out_rg, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Radius_of_gyration.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            rad_gyration = pd.read_csv(csv_out_rg)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            rad_gyration.to_excel(writer, sheet_name="Radius of gyration", index=False)

        plot_line(
            output_path=png_out,
            x_data=rad_gyration.iloc[:, 0],
            y_data=rad_gyration.iloc[:, 1:],
            x_axis_label="t / ps",
            y_axis_label=r"$R_{g}$(t)",
            show_legend=False,
            plot_size=(6, 6),
        )

    # ===== Gyration tensor
    csv_out_tensor = os.path.join(dir_out, "Etc", "Gyration_tensor_eigenvalues.csv")
    if not os.path.exists(csv_out_tensor) or override:
        print("Calculating gyration tensor eigenvalues...")

        eigenvalues = mdtraj_analysis.gyration_tensor(mdt_coords=mdt_resid_traj)

        results = {
            "Gyration tensor λ1 (average)": float(eigenvalues["n = 1"].mean()),
            "Gyration tensor λ2 (average)": float(eigenvalues["n = 2"].mean()),
            "Gyration tensor λ3 (average)": float(eigenvalues["n = 3"].mean()),
        }

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Radius of gyration",
            file_path=csv_out_tensor,
            parameters=params,
            results=results,
        )
        eigenvalues.to_csv(csv_out_tensor, index=False)

    # ===== Plot and save to .xlsx
    png_out = os.path.join(dir_out, "Etc", "Gyration_tensor_eigenvalues.png")
    if not os.path.exists(png_out) or override or plot_only:
        try:
            eigenvalues = pd.read_csv(csv_out_tensor)

        except FileNotFoundError:
            raise FileNotFoundError("CSV files must be generated prior to plotting.")

        with pd.ExcelWriter(
            xlsx_out, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            eigenvalues.to_excel(writer, sheet_name="Gyration tensor", index=False)

        plot_line(
            output_path=png_out,
            x_data=eigenvalues.iloc[:, 0],
            y_data=eigenvalues.iloc[:, 1:],
            x_axis_label="t / ps",
            y_axis_label=r"$\lambda_{n}$(t) / $nm^2$",
        )

    # ============================================
    # ===== Step 15: Animate residue in pore =====
    # ============================================
    os.makedirs(os.path.join(dir_out, "HBonds"), exist_ok=True)

    # ===== Animate residue in pore
    mp4_out = os.path.join(dir_out, "HBonds", f"{res_name}_animation.mp4")
    if not os.path.exists(mp4_out) or override:
        print("Animating residue in pore...")

        hbonds.animate_resname(
            mda_coords=mda_coords,
            res_name=res_name,
            pore_diameter=pore_diameter,
            mp4_out=mp4_out,
            n_res=20,
        )

        log_analysis_yaml(
            log_path=yaml_out,
            analysis_name="Residue animation",
            file_path=mp4_out,
            parameters=params,
        )


cli.add_command(Run)

if __name__ == "__main__":
    cli()
